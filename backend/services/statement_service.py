"""Generate downloadable statements from a user's transactions.

Three formats share one query path so a CSV download and an emailed PDF are
identical in content — only the rendering differs:

  build_statement_data(...)   ── shared shape: header + summary + transactions + categories
  render_csv(data)            ── bytes
  render_xlsx(data)           ── bytes
  render_pdf(data)            ── bytes

Reports cover a calendar month by default but the helper accepts any window.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from typing import Iterable

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from models.expense import Expense
from models.user import User


@dataclass
class StatementRow:
    when:     datetime
    name:     str
    category: str
    txn_type: str          # EXPENSE | INCOME
    amount:   float        # always positive
    method:   str | None
    location: str | None
    note:     str | None


@dataclass
class CategoryAgg:
    name:     str
    amount:   float
    percent:  float        # 0..100 share of expense or income side


@dataclass
class StatementData:
    user_name:     str
    user_email:    str
    period_label:  str          # "May 2026" or "1 May → 14 May 2026"
    period_start:  datetime
    period_end:    datetime
    rows:          list[StatementRow]
    total_income:  float
    total_expense: float
    net:           float
    expense_categories: list[CategoryAgg]
    income_categories:  list[CategoryAgg]


# ─── Data ─────────────────────────────────────────────────────────

def _month_window(year: int, month: int) -> tuple[datetime, datetime, str]:
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1)
    else:
        end = datetime(year, month + 1, 1)
    label = start.strftime("%B %Y")
    return start, end, label


async def build_statement_data(
    db: AsyncSession,
    user: User,
    *,
    start: datetime,
    end: datetime,
    period_label: str | None = None,
) -> StatementData:
    when_col = func.coalesce(Expense.txn_date, Expense.created_at)
    res = await db.execute(
        select(Expense).where(
            Expense.user_id == user.id,
            when_col >= start,
            when_col <  end,
        ).order_by(when_col.asc())
    )
    rows = [
        StatementRow(
            when=(e.txn_date or e.created_at),
            name=e.name, category=e.category, txn_type=e.txn_type,
            amount=e.amount,
            method=e.payment_method, location=e.location, note=e.note,
        )
        for e in res.scalars().all()
    ]

    total_expense = sum(r.amount for r in rows if r.txn_type == "EXPENSE")
    total_income  = sum(r.amount for r in rows if r.txn_type == "INCOME")

    exp_by_cat: dict[str, float] = {}
    inc_by_cat: dict[str, float] = {}
    for r in rows:
        m = inc_by_cat if r.txn_type == "INCOME" else exp_by_cat
        m[r.category] = m.get(r.category, 0.0) + r.amount

    def _aggs(by: dict[str, float], total: float) -> list[CategoryAgg]:
        return [
            CategoryAgg(
                name=name, amount=amt,
                percent=round(amt / total * 100, 1) if total else 0,
            )
            for name, amt in sorted(by.items(), key=lambda x: -x[1])
        ]

    return StatementData(
        user_name=user.full_name or "NkapSave user",
        user_email=user.email or "",
        period_label=period_label or
            f"{start.strftime('%-d %b')} → {(end - timedelta(seconds=1)).strftime('%-d %b %Y')}",
        period_start=start, period_end=end,
        rows=rows,
        total_income=total_income, total_expense=total_expense,
        net=total_income - total_expense,
        expense_categories=_aggs(exp_by_cat, total_expense),
        income_categories=_aggs(inc_by_cat, total_income),
    )


# ─── Renderers ────────────────────────────────────────────────────

def _fmt(n: float) -> str:
    return f"{n:,.0f}".replace(",", " ")


def render_csv(data: StatementData) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["NkapSave statement", data.user_name, data.user_email, data.period_label])
    w.writerow([])
    w.writerow(["Summary"])
    w.writerow(["Total income (XAF)",  _fmt(data.total_income)])
    w.writerow(["Total expense (XAF)", _fmt(data.total_expense)])
    w.writerow(["Net (XAF)",           _fmt(data.net)])
    w.writerow([])
    w.writerow(["Date", "Description", "Category", "Type",
                "Amount (XAF)", "Method", "Location", "Note"])
    for r in data.rows:
        w.writerow([
            r.when.strftime("%Y-%m-%d %H:%M"),
            r.name, r.category, r.txn_type,
            _fmt(r.amount),
            r.method or "", r.location or "", r.note or "",
        ])
    if data.expense_categories:
        w.writerow([])
        w.writerow(["Expense by category"])
        w.writerow(["Category", "Amount (XAF)", "% of expense"])
        for c in data.expense_categories:
            w.writerow([c.name, _fmt(c.amount), f"{c.percent:.1f}%"])
    if data.income_categories:
        w.writerow([])
        w.writerow(["Income by category"])
        w.writerow(["Category", "Amount (XAF)", "% of income"])
        for c in data.income_categories:
            w.writerow([c.name, _fmt(c.amount), f"{c.percent:.1f}%"])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel-on-Windows.


def render_xlsx(data: StatementData) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    title_font = Font(name="Calibri", size=16, bold=True, color="0E2A2A")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="12E8A4")
    label_font = Font(bold=True, color="555555")
    border_thin = Border(*[Side(style="thin", color="DDDDDD")] * 4)

    ws["A1"] = "NkapSave Statement"
    ws["A1"].font = title_font
    ws["A2"] = data.user_name
    ws["A3"] = data.user_email
    ws["A4"] = data.period_label

    ws["A6"] = "Summary";       ws["A6"].font = label_font
    ws["A7"] = "Total income";  ws["B7"] = data.total_income
    ws["A8"] = "Total expense"; ws["B8"] = data.total_expense
    ws["A9"] = "Net";           ws["B9"] = data.net
    for cell in ("B7", "B8", "B9"):
        ws[cell].number_format = '#,##0 "XAF"'

    headers = ["Date", "Description", "Category", "Type", "Amount",
               "Method", "Location", "Note"]
    ws.append([])
    ws.append(headers)
    hdr_row = ws.max_row
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")
        c.border = border_thin

    for r in data.rows:
        signed = r.amount if r.txn_type == "INCOME" else -r.amount
        ws.append([
            r.when.strftime("%Y-%m-%d %H:%M"),
            r.name, r.category, r.txn_type,
            signed,
            r.method or "", r.location or "", r.note or "",
        ])
        amount_cell = ws.cell(row=ws.max_row, column=5)
        amount_cell.number_format = '#,##0 "XAF";[Red]-#,##0 "XAF"'

    # Auto-ish column widths.
    widths = [18, 28, 14, 10, 14, 12, 18, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Category breakdown on second sheet.
    if data.expense_categories or data.income_categories:
        ws2 = wb.create_sheet(title="Breakdown")
        ws2.append(["Expense by category", "", "", "Income by category"])
        ws2.append(["Category", "Amount", "%",   "Category", "Amount", "%"])
        for c in ws2[2]:
            c.font = header_font
            c.fill = header_fill
        max_rows = max(len(data.expense_categories), len(data.income_categories))
        for i in range(max_rows):
            row = []
            if i < len(data.expense_categories):
                e = data.expense_categories[i]
                row.extend([e.name, e.amount, e.percent / 100])
            else:
                row.extend(["", "", ""])
            if i < len(data.income_categories):
                inc = data.income_categories[i]
                row.extend([inc.name, inc.amount, inc.percent / 100])
            else:
                row.extend(["", "", ""])
            ws2.append(row)
            ws2.cell(row=ws2.max_row, column=2).number_format = '#,##0 "XAF"'
            ws2.cell(row=ws2.max_row, column=3).number_format = "0.0%"
            ws2.cell(row=ws2.max_row, column=5).number_format = '#,##0 "XAF"'
            ws2.cell(row=ws2.max_row, column=6).number_format = "0.0%"
        for i, w in enumerate([22, 16, 8, 22, 16, 8], start=1):
            ws2.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_pdf(data: StatementData) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )

    buf = io.BytesIO()
    # PDF metadata (title/author) goes through PDFDocEncoding which is latin-1-ish
    # and rejects characters like → or em-dash. Strip them out of the metadata
    # only — the rendered body uses a real font and supports full Unicode.
    safe_label = (data.period_label
                  .replace("→", "to").replace("—", "-").replace("·", "-"))
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=f"NkapSave Statement - {safe_label}",
        author="NkapSave",
    )
    styles = getSampleStyleSheet()
    brand = colors.HexColor("#12E8A4")
    dark  = colors.HexColor("#0E2A2A")
    grey1 = colors.HexColor("#999999")
    grey2 = colors.HexColor("#DDDDDD")

    h1 = ParagraphStyle("h1", parent=styles["Heading1"],
        fontName="Helvetica-Bold", fontSize=20, textColor=dark, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["BodyText"],
        fontSize=10, textColor=grey1, spaceAfter=14)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"],
        fontName="Helvetica-Bold", fontSize=12, textColor=dark,
        spaceBefore=12, spaceAfter=6)
    body = ParagraphStyle("body", parent=styles["BodyText"],
        fontSize=9.5, leading=14)

    story = []
    story.append(Paragraph("NkapSave statement", h1))
    story.append(Paragraph(
        f"{data.user_name} &nbsp;·&nbsp; {data.user_email} &nbsp;·&nbsp; "
        f"<b>{data.period_label}</b>", sub,
    ))

    # Summary card as a 3-col table
    income  = _fmt(data.total_income)
    expense = _fmt(data.total_expense)
    net     = _fmt(data.net)
    summary = Table(
        [
            ["Total income", "Total expense", "Net"],
            [f"{income} XAF", f"{expense} XAF", f"{net} XAF"],
        ],
        colWidths=[55 * mm, 55 * mm, 55 * mm],
    )
    summary.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#F5F5F5")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), grey1),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("FONTNAME",    (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 1), (-1, 1), 14),
        ("TEXTCOLOR",   (0, 1), (0, 1), brand),
        ("TEXTCOLOR",   (1, 1), (1, 1), colors.HexColor("#FF4D6D")),
        ("TEXTCOLOR",   (2, 1), (2, 1),
            brand if data.net >= 0 else colors.HexColor("#FF4D6D")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOX",         (0, 0), (-1, -1), 0.5, grey2),
        ("GRID",        (0, 0), (-1, -1), 0.25, grey2),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(summary)

    # Categories
    def _cat_table(rows: list[CategoryAgg], title: str, color: colors.Color):
        story.append(Paragraph(title, h2))
        body_rows = [["Category", "Amount", "Share"]]
        for c in rows:
            body_rows.append([
                c.name, f"{_fmt(c.amount)} XAF", f"{c.percent:.1f}%",
            ])
        t = Table(body_rows, colWidths=[80 * mm, 50 * mm, 35 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), color),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 9),
            ("FONTSIZE",    (0, 1), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("GRID",        (0, 0), (-1, -1), 0.25, grey2),
            ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(t)

    if data.expense_categories:
        _cat_table(data.expense_categories,
                    "Expense by category",
                    colors.HexColor("#FF4D6D"))
    if data.income_categories:
        _cat_table(data.income_categories,
                    "Income by category",
                    brand)

    # Transaction list
    story.append(PageBreak())
    story.append(Paragraph("All transactions", h1))
    story.append(Paragraph(
        f"{len(data.rows)} entries · sorted oldest → newest", sub,
    ))
    headers = ["Date", "Description", "Category", "Method", "Amount"]
    tx_rows = [headers]
    for r in data.rows:
        sign = "+" if r.txn_type == "INCOME" else "−"
        tx_rows.append([
            r.when.strftime("%d %b"),
            r.name[:32] + ("…" if len(r.name) > 32 else ""),
            r.category,
            (r.method or "—")[:14],
            f"{sign}{_fmt(r.amount)}",
        ])
    if len(tx_rows) == 1:
        story.append(Paragraph(
            "No transactions in this period.", body,
        ))
    else:
        t = Table(tx_rows, colWidths=[20 * mm, 60 * mm, 30 * mm, 30 * mm, 30 * mm],
                   repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), dark),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 9),
            ("FONTSIZE",    (0, 1), (-1, -1), 8.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("GRID",        (0, 0), (-1, -1), 0.25, grey2),
            ("ALIGN",       (4, 0), (4, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                [colors.white, colors.HexColor("#FAFAFA")]),
        ]))
        # Colour income rows green, expense rows red on the amount cell.
        for i, r in enumerate(data.rows, start=1):
            color = brand if r.txn_type == "INCOME" else colors.HexColor("#FF4D6D")
            t.setStyle(TableStyle([("TEXTCOLOR", (4, i), (4, i), color)]))
        story.append(t)

    doc.build(story)
    return buf.getvalue()
